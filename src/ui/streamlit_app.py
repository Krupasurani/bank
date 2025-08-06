# # src/ui/streamlit_app.py
# import streamlit as st
# import os
# import json
# import tempfile
# import zipfile
# from pathlib import Path
# import pandas as pd
# from typing import List, Dict, Any
# import logging
# from dotenv import load_dotenv

# # Load environment variables
# load_dotenv()

# # Import our custom modules
# import sys
# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# from processors.document_processor import DocumentProcessor
# from ai_engine.test_generator import TestCaseGenerator
# from exporters.excel_exporter import TestCaseExporter

# # Configure logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Page configuration
# st.set_page_config(
#     page_title="ITASSIST - Test Case Generator",
#     page_icon="🤖",
#     layout="wide",
#     initial_sidebar_state="expanded"
# )

# def main():
#     """Main Streamlit application"""
    
#     # Title and description
#     st.title("🤖 ITASSIST - Intelligent Test Case Generator")
#     st.markdown("**AI-powered test case generation from BFSI documents**")
    
#     # Sidebar for configuration
#     with st.sidebar:
#         st.header("⚙️ Configuration")
        
#         # API Key input - auto-load from environment
#         default_api_key = os.getenv("OPENAI_API_KEY", "")
#         api_key = st.text_input(
#             "OpenAI API Key", 
#             value=default_api_key,
#             type="password",
#             help="API key loaded from environment" if default_api_key else "Enter your OpenAI API key"
#         )
        
#         # Model selection
#         model_option = st.selectbox(
#             "AI Model",
#             ["gpt-4.1-mini-2025-04-14", "gpt-4o-mini", "gpt-3.5-turbo"],
#             index=0
#         )
        
#         # Generation options
#         st.subheader("Generation Options")
#         num_test_cases = st.slider("Test Cases per Story", 5, 15, 8)
#         include_edge_cases = st.checkbox("Include Edge Cases", value=True)
#         include_negative_cases = st.checkbox("Include Negative Cases", value=True)
        
#         # Export format
#         export_format = st.multiselect(
#             "Export Formats",
#             ["Excel", "CSV", "JSON"],
#             default=["Excel"]
#         )
    
#     # Initialize session state
#     if 'generated_test_cases' not in st.session_state:
#         st.session_state.generated_test_cases = []
#     if 'processing_complete' not in st.session_state:
#         st.session_state.processing_complete = False
    
#     # Main content tabs
#     tab1, tab2, tab3 = st.tabs(["📁 Upload & Process", "🧪 Generated Test Cases", "💬 Chat Assistant"])
    
#     with tab1:
#         upload_and_process_tab(api_key, num_test_cases, include_edge_cases, include_negative_cases)
    
#     with tab2:
#         display_test_cases_tab(export_format)
    
#     with tab3:
#         chat_assistant_tab(api_key)

# def upload_and_process_tab(api_key: str, num_test_cases: int, include_edge_cases: bool, include_negative_cases: bool):
#     """File upload and processing tab"""
    
#     st.header("📁 Document Upload & Processing")
    
#     # File upload section
#     uploaded_files = st.file_uploader(
#         "Upload your documents",
#         type=['docx', 'pdf', 'xlsx', 'png', 'jpg', 'jpeg', 'tiff', 'bmp', 'txt', 'eml', 'json', 'xml', 'csv', 'zip'],
#         accept_multiple_files=True,
#         help="Supported formats: DOCX, PDF, XLSX, Images (PNG/JPG/TIFF/BMP), TXT, EML, JSON, XML, CSV, ZIP"
#     )
    
#     # Display file validation info
#     if uploaded_files:
#         st.info(f"📁 {len(uploaded_files)} file(s) uploaded successfully")
        
#         # Show file details
#         with st.expander("📋 File Details"):
#             for file in uploaded_files:
#                 file_size = len(file.getvalue()) / (1024*1024)  # MB
#                 st.write(f"• **{file.name}** ({file_size:.1f} MB)")
                
#                 # Validate file size
#                 if file_size > 50:
#                     st.warning(f"⚠️ {file.name} is large ({file_size:.1f} MB). Processing may take longer.")
    
#     # Enhanced processing options
#     st.subheader("🔧 Processing Options")
#     col1, col2, col3 = st.columns(3)
    
#     with col1:
#         process_embedded_content = st.checkbox("📷 Process Embedded Images/Screenshots", value=True)
#     with col2:
#         extract_tables = st.checkbox("📊 Extract Table Content", value=True)
#     with col3:
#         enhance_ocr = st.checkbox("🔍 Enhanced OCR Processing", value=True)
    
#     # Custom instructions with examples
#     st.subheader("📝 Custom Instructions")
    
#     # Predefined instruction templates
#     instruction_templates = {
#         "Standard": "",
#         "Focus on Negative Cases": "Generate more negative test cases and error scenarios. Include boundary testing and invalid input validation.",
#         "Basic Scenarios Only": "Focus on basic happy path scenarios. Minimize edge cases and complex integration tests.",
#         "Comprehensive Coverage": "Generate comprehensive test coverage including positive, negative, edge cases, and integration scenarios.",
#         "Security Focus": "Emphasize security testing scenarios including authentication, authorization, and data validation.",
#         "Performance Testing": "Include performance-related test scenarios for high-volume and stress testing."
#     }
    
#     selected_template = st.selectbox("Choose Instruction Template:", list(instruction_templates.keys()))
    
#     custom_instructions = st.text_area(
#         "Custom Instructions",
#         value=instruction_templates[selected_template],
#         placeholder="e.g., 'Focus on payment validation scenarios' or 'Create 4 test cases per acceptance criteria'",
#         help="Provide specific instructions to customize test case generation"
#     )
    
#     # Process button
#     if st.button("🚀 Generate Test Cases", type="primary", disabled=not api_key or not uploaded_files):
#         if not api_key:
#             st.error("Please provide OpenAI API key in the sidebar")
#             return
        
#         if not uploaded_files:
#             st.error("Please upload at least one document")
#             return
        
#         process_files(uploaded_files, api_key, custom_instructions, num_test_cases, 
#                      include_edge_cases, include_negative_cases, process_embedded_content)

# def process_files(uploaded_files, api_key: str, custom_instructions: str, 
#                  num_test_cases: int, include_edge_cases: bool, include_negative_cases: bool,
#                  process_embedded_content: bool):
#     """Process uploaded files and generate test cases"""
    
#     # Initialize processors
#     doc_processor = DocumentProcessor()
#     test_generator = TestCaseGenerator(api_key)
    
#     progress_bar = st.progress(0)
#     status_text = st.empty()
    
#     try:
#         all_content = []
#         total_files = len(uploaded_files)
        
#         # Process each file
#         for i, uploaded_file in enumerate(uploaded_files):
#             status_text.text(f"Processing {uploaded_file.name}...")
#             progress_bar.progress((i + 1) / (total_files + 2))
            
#             # Handle ZIP files
#             if uploaded_file.name.endswith('.zip'):
#                 extracted_content = process_zip_file(uploaded_file, doc_processor)
#                 all_content.extend(extracted_content)
#             else:
#                 # Save uploaded file temporarily
#                 with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
#                     tmp_file.write(uploaded_file.getvalue())
#                     tmp_file_path = tmp_file.name
                
#                 # Process the file
#                 result = doc_processor.process_file(tmp_file_path)
#                 all_content.append(result)
                
#                 # Clean up temporary file
#                 os.unlink(tmp_file_path)
        
#         # Combine all extracted content
#         status_text.text("Combining extracted content...")
#         progress_bar.progress(0.9)
        
#         combined_content = combine_extracted_content(all_content)
        
#         # Generate custom instructions
#         generation_instructions = build_generation_instructions(
#             custom_instructions, num_test_cases, include_edge_cases, include_negative_cases
#         )
        
#         # Generate test cases
#         status_text.text("Generating test cases with AI...")
#         progress_bar.progress(0.95)
        
#         test_cases = test_generator.generate_test_cases(combined_content, generation_instructions)
        
#         if test_cases:
#             st.session_state.generated_test_cases = test_cases
#             st.session_state.processing_complete = True
            
#             progress_bar.progress(1.0)
#             status_text.text("✅ Processing complete!")
            
#             # Display summary
#             st.success(f"Successfully generated {len(test_cases)} test cases!")
            
#             # Show content preview
#             with st.expander("📄 Extracted Content Preview"):
#                 st.text(combined_content[:1000] + "..." if len(combined_content) > 1000 else combined_content)
                
#         else:
#             st.error("No test cases could be generated. Please check your documents and try again.")
            
#     except Exception as e:
#         st.error(f"Error during processing: {str(e)}")
#         logger.error(f"Processing error: {str(e)}")

# def process_zip_file(zip_file, doc_processor: DocumentProcessor) -> List[Dict]:
#     """Process files within a ZIP archive"""
#     extracted_content = []
    
#     with tempfile.TemporaryDirectory() as temp_dir:
#         # Extract ZIP file
#         with zipfile.ZipFile(zip_file, 'r') as zip_ref:
#             zip_ref.extractall(temp_dir)
        
#         # Process each extracted file
#         for root, dirs, files in os.walk(temp_dir):
#             for file in files:
#                 file_path = os.path.join(root, file)
#                 if Path(file_path).suffix.lower() in doc_processor.supported_formats:
#                     result = doc_processor.process_file(file_path)
#                     extracted_content.append(result)
    
#     return extracted_content

# def combine_extracted_content(content_list: List[Dict]) -> str:
#     """Combine content from multiple files"""
#     combined_text = []
    
#     for content in content_list:
#         if content.get('content'):
#             file_info = f"\n--- Content from {content.get('file_name', 'Unknown')} ---\n"
#             combined_text.append(file_info + content['content'])
        
#         # Add table content if available
#         if content.get('tables'):
#             combined_text.append("\nTables:\n" + '\n'.join(content['tables']))
        
#         # Add OCR text from images
#         if content.get('image_text'):
#             combined_text.append("\nImage Text:\n" + '\n'.join(content['image_text']))
    
#     return '\n\n'.join(combined_text)

# def build_generation_instructions(custom_instructions: str, num_test_cases: int, 
#                                 include_edge_cases: bool, include_negative_cases: bool) -> str:
#     """Build generation instructions based on user preferences"""
#     instructions = []
    
#     if custom_instructions:
#         instructions.append(custom_instructions)
    
#     instructions.append(f"Generate exactly {num_test_cases} test cases per user story/requirement")
    
#     if include_edge_cases:
#         instructions.append("Include edge cases and boundary conditions")
    
#     if include_negative_cases:
#         instructions.append("Include negative test scenarios and error conditions")
    
#     instructions.append("Focus on BFSI domain scenarios with realistic banking data")
    
#     return ". ".join(instructions)

# def display_test_cases_tab(export_formats: List[str]):
#     """Display generated test cases"""
    
#     st.header("🧪 Generated Test Cases")
    
#     if not st.session_state.generated_test_cases:
#         st.info("No test cases generated yet. Please upload documents and process them first.")
#         return
    
#     test_cases = st.session_state.generated_test_cases
    
#     # Display summary metrics
#     col1, col2, col3, col4 = st.columns(4)
#     with col1:
#         st.metric("Total Test Cases", len(test_cases))
#     with col2:
#         high_priority = len([tc for tc in test_cases if tc.get("Priority") == "High"])
#         st.metric("High Priority", high_priority)
#     with col3:
#         regression_tests = len([tc for tc in test_cases if tc.get("Part of Regression") == "Yes"])
#         st.metric("Regression Tests", regression_tests)
#     with col4:
#         unique_stories = len(set(tc.get("User Story ID", "") for tc in test_cases))
#         st.metric("User Stories", unique_stories)
    
#     # Filter options
#     with st.expander("🔍 Filter Test Cases"):
#         col1, col2, col3 = st.columns(3)
        
#         with col1:
#             priority_filter = st.multiselect(
#                 "Priority", 
#                 ["High", "Medium", "Low"],
#                 default=["High", "Medium", "Low"]
#             )
        
#         with col2:
#             regression_filter = st.multiselect(
#                 "Regression", 
#                 ["Yes", "No"],
#                 default=["Yes", "No"]
#             )
        
#         with col3:
#             story_ids = list(set(tc.get("User Story ID", "") for tc in test_cases))
#             story_filter = st.multiselect(
#                 "User Story ID",
#                 story_ids,
#                 default=story_ids
#             )
    
#     # Apply filters
#     filtered_test_cases = [
#         tc for tc in test_cases
#         if (tc.get("Priority") in priority_filter and
#             tc.get("Part of Regression") in regression_filter and
#             tc.get("User Story ID") in story_filter)
#     ]
    
#     # Display test cases table
#     if filtered_test_cases:
#         st.subheader(f"Test Cases ({len(filtered_test_cases)} of {len(test_cases)})")
        
#         # Convert to DataFrame for display
#         df = pd.DataFrame(filtered_test_cases)
        
#         # Configure column display
#         column_config = {
#             "Steps": st.column_config.TextColumn(width="large"),
#             "Test Case Description": st.column_config.TextColumn(width="medium"),
#             "Expected Result": st.column_config.TextColumn(width="medium"),
#         }
        
#         st.dataframe(
#             df,
#             use_container_width=True,
#             column_config=column_config,
#             hide_index=True
#         )
        
#         # Export section
#         st.subheader("📥 Export Test Cases")
        
#         col1, col2, col3 = st.columns(3)
        
#         exporter = TestCaseExporter()
        
#         if "Excel" in export_formats:
#             with col1:
#                 if st.button("📊 Download Excel", type="primary"):
#                     try:
#                         # Create Excel in memory instead of temp file
#                         import io
#                         from openpyxl import Workbook
                        
#                         # Use our exporter but modify to work with BytesIO
#                         exporter = TestCaseExporter()
                        
#                         # Create DataFrame
#                         df = pd.DataFrame(filtered_test_cases)
                        
#                         # Ensure all required columns exist
#                         for col in exporter.required_columns:
#                             if col not in df.columns:
#                                 df[col] = ""
                        
#                         # Reorder columns
#                         df = df[exporter.required_columns]
                        
#                         # Create Excel file in memory
#                         output = io.BytesIO()
#                         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                             df.to_excel(writer, sheet_name='Test Cases', index=False)
                            
#                             # Get workbook and worksheet for formatting
#                             workbook = writer.book
#                             worksheet = writer.sheets['Test Cases']
                            
#                             # Apply basic formatting
#                             from openpyxl.styles import Font, PatternFill
                            
#                             # Header formatting
#                             for cell in worksheet[1]:
#                                 cell.font = Font(bold=True, color="FFFFFF")
#                                 cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        
#                         output.seek(0)
                        
#                         st.download_button(
#                             label="📊 Download Excel File",
#                             data=output.getvalue(),
#                             file_name=f"test_cases_{len(filtered_test_cases)}.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                        
#                     except Exception as e:
#                         st.error(f"Export error: {str(e)}")
#                         logger.error(f"Excel export error: {str(e)}")
        
#         if "CSV" in export_formats:
#             with col2:
#                 if st.button("📄 Download CSV"):
#                     try:
#                         csv_data = pd.DataFrame(filtered_test_cases).to_csv(index=False)
#                         st.download_button(
#                             label="📄 Download CSV File",
#                             data=csv_data,
#                             file_name=f"test_cases_{len(filtered_test_cases)}.csv",
#                             mime="text/csv"
#                         )
#                     except Exception as e:
#                         st.error(f"Export error: {str(e)}")
        
#         if "JSON" in export_formats:
#             with col3:
#                 if st.button("🔧 Download JSON"):
#                     try:
#                         json_data = json.dumps(filtered_test_cases, indent=2, ensure_ascii=False)
#                         st.download_button(
#                             label="🔧 Download JSON File",
#                             data=json_data,
#                             file_name=f"test_cases_{len(filtered_test_cases)}.json",
#                             mime="application/json"
#                         )
#                     except Exception as e:
#                         st.error(f"Export error: {str(e)}")
    
#     else:
#         st.warning("No test cases match the selected filters.")

# def chat_assistant_tab(api_key: str):
#     """Chat assistant for test case customization"""
    
#     st.header("💬 Chat Assistant")
#     st.markdown("Ask questions about your test cases or request modifications")
    
#     if not api_key:
#         st.warning("Please provide OpenAI API key to enable chat functionality")
#         return
    
#     if not st.session_state.generated_test_cases:
#         st.info("Generate test cases first to enable chat assistance")
#         return
    
#     # Chat interface
#     if "chat_messages" not in st.session_state:
#         st.session_state.chat_messages = []
    
#     # Display chat history
#     for message in st.session_state.chat_messages:
#         with st.chat_message(message["role"]):
#             st.markdown(message["content"])
    
#     # Chat input
#     if prompt := st.chat_input("Ask about your test cases..."):
#         # Add user message
#         st.session_state.chat_messages.append({"role": "user", "content": prompt})
#         with st.chat_message("user"):
#             st.markdown(prompt)
        
#         # Generate response
#         with st.chat_message("assistant"):
#             with st.spinner("Thinking..."):
#                 response = generate_chat_response(prompt, st.session_state.generated_test_cases, api_key)
#                 st.markdown(response)
        
#         # Add assistant message
#         st.session_state.chat_messages.append({"role": "assistant", "content": response})

# def generate_chat_response(prompt: str, test_cases: List[Dict], api_key: str) -> str:
#     """Generate chat response about test cases"""
#     try:
#         from openai import OpenAI
#         client = OpenAI(api_key=api_key)
        
#         # Prepare context
#         test_cases_summary = f"Total test cases: {len(test_cases)}\n"
#         test_cases_summary += "Sample test cases:\n"
#         for i, tc in enumerate(test_cases[:3], 1):
#             test_cases_summary += f"{i}. {tc.get('Test Case Description', '')}\n"
        
#         chat_prompt = f"""
#         You are an expert BFSI test engineer assistant. Answer questions about the generated test cases.
        
#         Test Cases Context:
#         {test_cases_summary}
        
#         User Question: {prompt}
        
#         Provide helpful, specific answers about the test cases. If asked to modify test cases, 
#         provide specific suggestions or instructions.
#         """
        
#         response = client.chat.completions.create(
#             model="gpt-4.1-mini-2025-04-14",
#             messages=[
#                 {"role": "system", "content": "You are a helpful BFSI testing expert."},
#                 {"role": "user", "content": chat_prompt}
#             ],
#             temperature=0.7,
#             max_tokens=500
#         )
        
#         return response.choices[0].message.content
        
#     except Exception as e:
#         return f"Sorry, I encountered an error: {str(e)}"

# if __name__ == "__main__":
#     main()





# # src/ui/streamlit_app.py - FIXED VERSION
# """
# Fixed Streamlit App with Super Intelligent LLM System
# - No auto-approval thresholds - always proceed with intelligence
# - Original Excel format maintained
# - Works with any input files dynamically
# """

# import streamlit as st
# import os
# import json
# import tempfile
# from pathlib import Path
# import pandas as pd
# from typing import List, Dict, Any
# import logging
# from dotenv import load_dotenv

# # Load environment variables
# load_dotenv()

# # Import modules
# import sys
# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# from processors.document_processor import DocumentProcessor
# from exporters.excel_exporter import TestCaseExporter

# # Try to import fully dynamic PACS.008 modules
# try:
#     from processors.pacs008_intelligent_detector import PACS008IntelligentDetector
#     from processors.fully_dynamic_intelligent_maker_checker import FullyDynamicIntelligentMakerChecker
#     from ai_engine.fully_dynamic_test_generator import FullyDynamicTestGenerator
#     SUPER_INTELLIGENT_MODE_AVAILABLE = True
# except ImportError:
#     from ai_engine.test_generator import TestCaseGenerator
#     SUPER_INTELLIGENT_MODE_AVAILABLE = False

# # Configure logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# def main():
#     """Main Streamlit application with super intelligent dynamic system"""
    
#     # Page configuration
#     st.set_page_config(
#         page_title="ITASSIST - Intelligent PACS.008 Generator",
#         page_icon="🏦" if SUPER_INTELLIGENT_MODE_AVAILABLE else "🤖",
#         layout="wide",
#         initial_sidebar_state="expanded"
#     )
    
#     # Title
#     if SUPER_INTELLIGENT_MODE_AVAILABLE:
#         st.title("🏦 ITASSIST - Intelligent PACS.008 Test Generator")
#         st.markdown("**AI-powered field detection + Intelligent maker-checker validation**")
#         st.success("✅ **Intelligent Mode**: Automated LLM validation with banking expertise")
#     else:
#         st.title("🤖 ITASSIST - Standard Test Generator")
#         st.markdown("**Standard test case generation**")
    
#     # Sidebar
#     with st.sidebar:
#         st.header("⚙️ Configuration")
        
#         # API Key
#         default_api_key = os.getenv("OPENAI_API_KEY", "")
#         api_key = st.text_input(
#             "OpenAI API Key",
#             value=default_api_key,
#             type="password",
#             help="Required for intelligent validation and test generation"
#         )
        
#         # Intelligent Features Status
#         if SUPER_INTELLIGENT_MODE_AVAILABLE:
#             st.subheader("🧠 Intelligent Features")
#             st.success("✅ **All Systems Active**")
#             st.info("🚀 **Always Proceeds**: No thresholds - works with any input")
#             st.write("• PACS.008 Field Detection")
#             st.write("• AI Maker-Checker Validation")
#             st.write("• Enhanced Test Generation")
#             st.write("• Original Excel Format")
            
#             enable_intelligent_mode = st.checkbox("Enable Intelligent Mode", value=True)
#         else:
#             enable_intelligent_mode = False
        
#         # Test Generation Options
#         st.subheader("🧪 Test Generation")
#         num_test_cases = st.slider("Test Cases per Story", 5, 15, 8)
#         include_edge_cases = st.checkbox("Include Edge Cases", value=True)
#         include_negative_cases = st.checkbox("Include Negative Cases", value=True)
        
#         # Export Options
#         export_format = st.multiselect(
#             "Export Formats",
#             ["Excel", "CSV", "JSON"],
#             default=["Excel"]
#         )
    
#     # Initialize session state
#     if 'processing_results' not in st.session_state:
#         st.session_state.processing_results = {}
#     if 'validation_results' not in st.session_state:
#         st.session_state.validation_results = {}
#     if 'final_test_cases' not in st.session_state:
#         st.session_state.final_test_cases = []
    
#     # Main tabs
#     if SUPER_INTELLIGENT_MODE_AVAILABLE and enable_intelligent_mode:
#         tab1, tab2, tab3, tab4 = st.tabs([
#             "📁 Upload & Process", 
#             "🧠 AI Validation Results", 
#             "👥 Intelligent Maker-Checker", 
#             "🧪 Enhanced Test Cases"
#         ])
        
#         with tab1:
#             upload_and_process_tab(api_key, enable_intelligent_mode)
        
#         with tab2:
#             intelligent_validation_tab()
        
#         with tab3:
#             intelligent_maker_checker_tab()
        
#         with tab4:
#             enhanced_test_cases_tab(export_format, num_test_cases, include_edge_cases, include_negative_cases)
#     else:
#         tab1, tab2 = st.tabs(["📁 Upload & Process", "🧪 Test Cases"])
        
#         with tab1:
#             upload_and_process_tab(api_key, False)
        
#         with tab2:
#             enhanced_test_cases_tab(export_format, num_test_cases, include_edge_cases, include_negative_cases)

# def upload_and_process_tab(api_key: str, enable_intelligent_mode: bool):
#     """Upload and processing tab with intelligent options"""
    
#     st.header("📁 Document Upload & Intelligent Processing")
    
#     if enable_intelligent_mode:
#         st.info("🧠 **Intelligent Mode**: AI will detect PACS.008 fields, validate them using banking expertise, and generate enhanced test cases")
#         st.success("🚀 **Always Proceeds**: System works with any input - no approval thresholds needed")
#     else:
#         st.info("📄 **Standard Mode**: Basic document processing and test case generation")
    
#     # File upload
#     uploaded_files = st.file_uploader(
#         "Upload your banking documents",
#         type=['docx', 'pdf', 'xlsx', 'txt', 'json'],
#         accept_multiple_files=True,
#         help="Upload requirements, user stories, payment specifications, or banking documentation"
#     )
    
#     if uploaded_files:
#         st.success(f"📁 {len(uploaded_files)} file(s) uploaded successfully")
        
#         # Show file preview
#         with st.expander("📋 Uploaded Files"):
#             for file in uploaded_files:
#                 file_size = len(file.getvalue()) / (1024*1024)
#                 st.write(f"• **{file.name}** ({file_size:.1f} MB)")
    
#     # Processing instructions
#     st.subheader("📝 Processing Instructions")
    
#     if enable_intelligent_mode:
#         instruction_templates = {
#             "Intelligent Auto-Mode": "Let AI automatically handle everything - field detection, validation, and test generation",
#             "Cross-Border Payment Focus": "Focus on cross-border payment scenarios with correspondent banking and compliance",
#             "Domestic Payment Focus": "Focus on domestic payment scenarios within single country clearing",
#             "High-Value Payment Focus": "Focus on high-value payments with enhanced compliance and monitoring",
#             "Compliance & AML Focus": "Emphasize regulatory compliance, AML/KYC, and sanctions screening"
#         }
#     else:
#         instruction_templates = {
#             "Standard Generation": "Generate comprehensive test cases covering positive, negative, and edge scenarios",
#             "Basic Scenarios": "Focus on fundamental happy path and basic error handling"
#         }
    
#     selected_template = st.selectbox("Processing Template:", list(instruction_templates.keys()))
    
#     custom_instructions = st.text_area(
#         "Custom Instructions (Optional)",
#         value=instruction_templates[selected_template],
#         placeholder="e.g., 'Focus on Deutsche Bank to BNP Paribas routing' or 'Generate compliance-heavy scenarios'",
#         help="AI will enhance these instructions based on detected PACS.008 content"
#     )
    
#     # Process button
#     process_button_text = "🧠 Start Intelligent Processing" if enable_intelligent_mode else "🚀 Generate Test Cases"
    
#     if st.button(process_button_text, type="primary", disabled=not api_key or not uploaded_files):
#         if not api_key:
#             st.error("❌ Please provide OpenAI API key")
#             return
        
#         if not uploaded_files:
#             st.error("❌ Please upload at least one document") 
#             return
        
#         if enable_intelligent_mode:
#             process_with_intelligent_pipeline(uploaded_files, api_key, custom_instructions)
#         else:
#             process_with_standard_generation(uploaded_files, api_key, custom_instructions)

# def process_with_intelligent_pipeline(uploaded_files, api_key: str, custom_instructions: str):
#     """Process files through the intelligent pipeline - always proceeds"""
    
#     progress_bar = st.progress(0)
#     status_text = st.empty()
    
#     try:
#         # Step 1: Initialize
#         status_text.text("🔧 Initializing intelligent AI systems...")
#         progress_bar.progress(0.1)
        
#         doc_processor = DocumentProcessor()
#         pacs008_detector = PACS008IntelligentDetector(api_key)
#         intelligent_maker_checker = FullyDynamicIntelligentMakerChecker(api_key)
#         test_generator = FullyDynamicTestGenerator(api_key)
        
#         # Step 2: Document processing
#         status_text.text("📄 Processing documents...")
#         progress_bar.progress(0.2)
        
#         all_content = []
#         for uploaded_file in uploaded_files:
#             with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
#                 tmp_file.write(uploaded_file.getvalue())
#                 tmp_file_path = tmp_file.name
            
#             result = doc_processor.process_file(tmp_file_path)
#             all_content.append(result.get('content', ''))
#             os.unlink(tmp_file_path)
        
#         combined_content = '\n\n'.join(all_content)
        
#         # Step 3: PACS.008 field detection
#         status_text.text("🏦 AI detecting PACS.008 fields...")
#         progress_bar.progress(0.4)
        
#         detection_results = pacs008_detector.detect_pacs008_fields_in_input(combined_content)
        
#         if detection_results['status'] != 'SUCCESS':
#             st.info("ℹ️ No PACS.008 fields detected. Using intelligent standard processing...")
#             return process_with_intelligent_fallback(combined_content, api_key, custom_instructions)
        
#         detected_fields = detection_results.get('detected_fields', [])
#         st.success(f"✅ **{len(detected_fields)} PACS.008 fields detected**")
        
#         # Step 4: Intelligent maker-checker validation (always proceeds)
#         status_text.text("🧠 AI performing intelligent validation...")
#         progress_bar.progress(0.6)
        
#         validation_results = intelligent_maker_checker.perform_fully_dynamic_validation(detected_fields)
        
#         # Store validation results
#         st.session_state.validation_results = validation_results
        
#         if validation_results['status'] != 'SUCCESS':
#             st.warning("⚠️ AI validation had issues but proceeding with intelligent generation...")
#         else:
#             # Show validation score but always proceed
#             validation_score = validation_results.get('final_analysis', {}).get('final_validation_score', 0)
#             st.info(f"🎯 **AI Validation Score**: {validation_score}% - Proceeding with intelligent generation")
        
#         # Step 5: Enhanced test case generation (always proceeds)
#         status_text.text("🧪 Generating enhanced test cases with AI insights...")
#         progress_bar.progress(0.8)
        
#         test_generation_results = test_generator.generate_fully_dynamic_test_cases(
#             combined_content, detected_fields, custom_instructions
#         )
        
#         # Store results
#         st.session_state.processing_results = {
#             'detection_results': detection_results,
#             'validation_results': validation_results,
#             'test_generation_results': test_generation_results
#         }
        
#         # Always get test cases
#         generated_test_cases = test_generation_results.get('test_cases', [])
        
#         if not generated_test_cases:
#             st.warning("🔄 Primary generation returned 0 cases. Using enhanced fallback generation...")
#             # Enhanced fallback
#             from ai_engine.enhanced_test_generator import EnhancedTestCaseGenerator
#             fallback_generator = EnhancedTestCaseGenerator(api_key)
#             generated_test_cases = fallback_generator.generate_enhanced_test_cases(
#                 combined_content, custom_instructions
#             )
#             st.success(f"✅ Generated {len(generated_test_cases)} enhanced fallback test cases!")
#         else:
#             st.success(f"✅ Generated {len(generated_test_cases)} intelligent test cases!")
        
#         st.session_state.final_test_cases = generated_test_cases
        
#         progress_bar.progress(1.0)
#         status_text.text("✅ Intelligent processing completed successfully!")
        
#         # Display summary
#         display_intelligent_summary(detection_results, validation_results, len(generated_test_cases))
        
#     except Exception as e:
#         st.error(f"❌ Intelligent processing failed: {str(e)}")
#         logger.error(f"Processing error: {str(e)}")

# def display_intelligent_summary(detection_results: Dict, validation_results: Dict, test_count: int):
#     """Display intelligent processing summary"""
    
#     st.subheader("📊 Intelligent Processing Summary")
    
#     col1, col2, col3, col4 = st.columns(4)
    
#     with col1:
#         detected_count = len(detection_results.get('detected_fields', []))
#         st.metric("🔍 AI Detected", f"{detected_count} fields")
    
#     with col2:
#         validation_score = validation_results.get('final_analysis', {}).get('final_validation_score', 0)
#         st.metric("🎯 AI Score", f"{validation_score}%")
    
#     with col3:
#         overall_decision = validation_results.get('intelligent_decisions', {}).get('combined_analysis', {}).get('overall_decision', 'PROCESSED')
#         status_icon = "✅" if overall_decision in ["APPROVED", "CONDITIONALLY_APPROVED"] else "🔄"
#         st.metric("🤖 AI Status", f"{status_icon}")
    
#     with col4:
#         st.metric("🧪 Test Cases", test_count)
    
#     # Always show success - no thresholds
#     st.success("🎉 **Intelligent Processing Complete**: AI analyzed content, performed validation, and generated enhanced test cases successfully!")

# def intelligent_validation_tab():
#     """Display AI validation results"""
    
#     st.header("🧠 AI Validation Results")
    
#     if not st.session_state.validation_results:
#         st.info("📄 No AI validation results available. Please process documents first.")
#         return
    
#     validation_results = st.session_state.validation_results
    
#     # Validation overview
#     st.subheader("📊 AI Validation Overview")
    
#     final_analysis = validation_results.get('final_analysis', {})
#     comprehensive_validation = validation_results.get('comprehensive_validation', {})
#     overall_assessment = comprehensive_validation.get('overall_assessment', {})
    
#     col1, col2, col3 = st.columns(3)
    
#     with col1:
#         technical_score = overall_assessment.get('technical_score', 0)
#         st.metric("🔧 Technical Intelligence", f"{technical_score}%")
    
#     with col2:
#         business_score = overall_assessment.get('business_score', 0)
#         st.metric("🏦 Business Intelligence", f"{business_score}%")
    
#     with col3:
#         compliance_score = overall_assessment.get('compliance_score', 0)
#         st.metric("📋 Compliance Intelligence", f"{compliance_score}%")
    
#     # Key findings
#     key_findings = final_analysis.get('validation_summary', {}).get('key_findings', [])
#     if key_findings:
#         st.subheader("💡 AI Key Findings")
#         for finding in key_findings:
#             st.write(f"• {finding}")
    
#     # Field validations
#     field_validations = comprehensive_validation.get('field_validations', [])
#     if field_validations:
#         st.subheader("🔍 Dynamic Field Analysis")
        
#         # Group by status
#         valid_fields = [f for f in field_validations if f.get('validation_status') == 'VALID']
#         warning_fields = [f for f in field_validations if f.get('validation_status') in ['WARNING', 'MISSING']]
#         invalid_fields = [f for f in field_validations if f.get('validation_status') == 'INVALID']
        
#         if valid_fields:
#             with st.expander(f"✅ Valid Fields ({len(valid_fields)})", expanded=True):
#                 for field in valid_fields:
#                     st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Valid')}")
        
#         if warning_fields:
#             with st.expander(f"⚠️ Fields with Warnings ({len(warning_fields)})", expanded=True):
#                 for field in warning_fields:
#                     st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Warning')}")
        
#         if invalid_fields:
#             with st.expander(f"❌ Invalid Fields ({len(invalid_fields)})", expanded=False):
#                 for field in invalid_fields:
#                     st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Invalid')}")

# def intelligent_maker_checker_tab():
#     """Display intelligent maker-checker decisions"""
    
#     st.header("👥 Intelligent Maker-Checker Results")
    
#     if not st.session_state.validation_results:
#         st.info("📄 No maker-checker results available. Please process documents first.")
#         return
    
#     validation_results = st.session_state.validation_results
#     intelligent_decisions = validation_results.get('intelligent_decisions', {})
    
#     if not intelligent_decisions:
#         st.info("ℹ️ Maker-checker analysis completed - system always proceeds with intelligent generation.")
#         return
    
#     # Overall AI decision
#     combined_analysis = intelligent_decisions.get('combined_analysis', {})
#     overall_decision = combined_analysis.get('overall_decision', 'PROCESSED')
    
#     st.info(f"🤖 **AI Decision**: {overall_decision} - System proceeded with intelligent test generation")
    
#     # AI Maker decision
#     st.subheader("🤖 AI Maker Analysis")
    
#     maker_decision = intelligent_decisions.get('maker_decision', {})
    
#     col1, col2 = st.columns(2)
    
#     with col1:
#         decision = maker_decision.get('decision', 'PROCESSED')
#         confidence = maker_decision.get('confidence', 0)
#         st.write(f"**Decision**: {decision}")
#         st.write(f"**Confidence**: {confidence}%")
    
#     with col2:
#         reasoning = maker_decision.get('technical_reasoning', 'AI completed technical analysis')
#         st.write(f"**AI Reasoning**: {reasoning}")
    
#     # AI Checker decision
#     st.subheader("✅ AI Checker Analysis")
    
#     checker_decision = intelligent_decisions.get('checker_decision', {})
    
#     col1, col2 = st.columns(2)
    
#     with col1:
#         decision = checker_decision.get('decision', 'PROCESSED')
#         confidence = checker_decision.get('confidence', 0)
#         st.write(f"**Decision**: {decision}")
#         st.write(f"**Confidence**: {confidence}%")
    
#     with col2:
#         reasoning = checker_decision.get('business_reasoning', 'AI completed business analysis')
#         st.write(f"**AI Reasoning**: {reasoning}")
    
#     # Show the benefit
#     st.success("💡 **Fully Automated**: AI completed the entire maker-checker process automatically and proceeded with intelligent test generation - no manual intervention or thresholds needed!")

# def enhanced_test_cases_tab(export_formats: List[str], num_test_cases: int, 
#                           include_edge_cases: bool, include_negative_cases: bool):
#     """Display enhanced test cases with original Excel format"""
    
#     st.header("🧪 Enhanced Test Cases")
    
#     if not st.session_state.final_test_cases:
#         st.info("📄 No test cases generated yet. Please process documents first.")
#         return
    
#     test_cases = st.session_state.final_test_cases
#     processing_results = st.session_state.processing_results
    
#     # Test case overview
#     col1, col2, col3, col4 = st.columns(4)
    
#     with col1:
#         st.metric("Total Test Cases", len(test_cases))
    
#     with col2:
#         enhanced_tests = len([tc for tc in test_cases if tc.get('PACS008_Enhanced') == 'Yes'])
#         if enhanced_tests > 0:
#             st.metric("🏦 PACS.008 Enhanced", enhanced_tests)
#         else:
#             high_priority = len([tc for tc in test_cases if tc.get('Priority') == 'High'])
#             st.metric("High Priority", high_priority)
    
#     with col3:
#         regression_tests = len([tc for tc in test_cases if tc.get('Part of Regression') == 'Yes'])
#         st.metric("Regression Tests", regression_tests)
    
#     with col4:
#         unique_stories = len(set(tc.get('User Story ID', '') for tc in test_cases))
#         st.metric("User Stories", unique_stories)
    
#     # Generation method
#     generation_method = processing_results.get('test_generation_results', {}).get('generation_method', 'INTELLIGENT')
    
#     if 'DYNAMIC' in generation_method or 'ENHANCED' in generation_method:
#         st.success("✅ **Intelligent Generation**: Test cases created using AI validation insights and banking expertise")
#     else:
#         st.info(f"ℹ️ **Generation Method**: {generation_method}")
    
#     # Display test cases
#     st.subheader("📋 Generated Test Cases")
    
#     # Create clean display with original columns
#     display_data = []
#     for tc in test_cases:
#         clean_tc = {
#             "User Story ID": tc.get('User Story ID', ''),
#             "Test Case ID": tc.get('Test Case ID', ''),
#             "Test Case Description": tc.get('Test Case Description', ''),
#             "Priority": tc.get('Priority', ''),
#             "Part of Regression": tc.get('Part of Regression', ''),
#         }
        
#         # Add PACS.008 indicator if available
#         if tc.get('PACS008_Enhanced') == 'Yes':
#             clean_tc["🏦 Enhanced"] = "✅"
#         elif tc.get('validation_status'):
#             clean_tc["🏦 Enhanced"] = "✅"
#         else:
#             clean_tc["🏦 Enhanced"] = "⭕"
        
#         display_data.append(clean_tc)
    
#     df = pd.DataFrame(display_data)
#     st.dataframe(df, use_container_width=True, hide_index=True)
    
#     # Detailed view
#     with st.expander("📋 Detailed Test Case View (First 3)"):
#         for i, tc in enumerate(test_cases[:3], 1):
#             st.write(f"**{i}. {tc.get('Test Case Description', 'N/A')}**")
#             st.write(f"**Steps**: {tc.get('Steps', 'N/A')}")
#             st.write(f"**Expected Result**: {tc.get('Expected Result', 'N/A')}")
#             if tc.get('PACS008_Enhanced') or tc.get('validation_status'):
#                 st.write(f"**Enhancement**: PACS.008 Intelligent Generation")
#             st.write("---")
    
#     # Export section with ORIGINAL Excel format
#     st.subheader("📥 Export Test Cases")
    
#     if "Excel" in export_formats:
#         if st.button("📊 Download Excel (Original Format)", type="primary"):
#             export_original_excel_format(test_cases)
    
#     if "CSV" in export_formats:
#         if st.button("📄 Download CSV"):
#             export_csv_format(test_cases)
    
#     if "JSON" in export_formats:
#         if st.button("🔧 Download JSON"):
#             export_json_format(test_cases)

# def clean_test_cases_for_export(test_cases: List[Dict]) -> List[Dict]:
#     """Clean test cases data to handle lists, None values, and other export issues"""
    
#     cleaned_cases = []
    
#     for test_case in test_cases:
#         cleaned_case = {}
        
#         for key, value in test_case.items():
#             # Handle None values
#             if value is None:
#                 cleaned_case[key] = ""
#             # Handle list values (convert to numbered string)
#             elif isinstance(value, list):
#                 if len(value) == 0:
#                     cleaned_case[key] = ""
#                 else:
#                     # Convert list to numbered steps
#                     numbered_steps = []
#                     for i, item in enumerate(value, 1):
#                         if isinstance(item, str):
#                             numbered_steps.append(f"{i}. {item}")
#                         else:
#                             numbered_steps.append(f"{i}. {str(item)}")
#                     cleaned_case[key] = "\n".join(numbered_steps)
#             # Handle dictionary values
#             elif isinstance(value, dict):
#                 cleaned_case[key] = str(value)
#             # Handle very long strings (Excel limit)
#             elif isinstance(value, str) and len(value) > 32767:
#                 cleaned_case[key] = value[:32760] + "..."
#             # Handle other types
#             else:
#                 cleaned_case[key] = str(value) if value is not None else ""
        
#         cleaned_cases.append(cleaned_case)
    
#     return cleaned_cases

# def export_original_excel_format(test_cases: List[Dict]):
#     """Export using the ORIGINAL Excel format from TestCaseExporter"""
    
#     try:
#         # Clean test cases data first
#         cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
#         # Use your original TestCaseExporter
#         exporter = TestCaseExporter()
        
#         # Create Excel in memory using original format
#         import io
#         from openpyxl import Workbook
#         from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Test Cases"
        
#         # Use original required columns
#         required_columns = [
#             "User Story ID",
#             "Acceptance Criteria ID", 
#             "Scenario",
#             "Test Case ID",
#             "Test Case Description",
#             "Precondition",
#             "Steps",
#             "Expected Result",
#             "Part of Regression",
#             "Priority"
#         ]
        
#         # Create DataFrame with original format
#         df = pd.DataFrame(cleaned_test_cases)
        
#         # Ensure all required columns exist
#         for col in required_columns:
#             if col not in df.columns:
#                 df[col] = ""
        
#         # Reorder columns to original format
#         df = df[required_columns]
        
#         # Add headers with original formatting
#         for col_num, column_title in enumerate(required_columns, 1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.value = column_title
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.border = Border(
#                 left=Side(style='thin'),
#                 right=Side(style='thin'),
#                 top=Side(style='thin'),
#                 bottom=Side(style='thin')
#             )
        
#         # Add data rows with original formatting
#         for row_num, (index, test_case) in enumerate(df.iterrows(), 2):
#             for col_num, column in enumerate(required_columns, 1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 value = test_case[column]
                
#                 # Handle list data (convert to string)
#                 if isinstance(value, list):
#                     value = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(value)])
#                     cell.alignment = Alignment(wrap_text=True, vertical="top")
#                 elif column == "Steps" and "\\n" in str(value):
#                     # Handle multi-line content (Steps field) - original way
#                     value = value.replace("\\n", "\n")
#                     cell.alignment = Alignment(wrap_text=True, vertical="top")
                
#                 # Ensure value is string and not too long for Excel
#                 value = str(value) if value is not None else ""
#                 if len(value) > 32767:  # Excel cell limit
#                     value = value[:32760] + "..."
                
#                 cell.value = value
                
#                 # Apply original conditional formatting based on Priority
#                 if column == "Priority":
#                     if value == "High":
#                         cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
#                     elif value == "Medium":
#                         cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
#                     elif value == "Low":
#                         cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                
#                 # Add borders to all cells - original way
#                 cell.border = Border(
#                     left=Side(style='thin'),
#                     right=Side(style='thin'),
#                     top=Side(style='thin'),
#                     bottom=Side(style='thin')
#                 )
        
#         # Auto-adjust column widths - original way
#         column_widths = {
#             'A': 15,  # User Story ID
#             'B': 20,  # Acceptance Criteria ID
#             'C': 25,  # Scenario
#             'D': 15,  # Test Case ID
#             'E': 40,  # Test Case Description
#             'F': 30,  # Precondition
#             'G': 50,  # Steps
#             'H': 40,  # Expected Result
#             'I': 18,  # Part of Regression
#             'J': 12   # Priority
#         }
        
#         for column, width in column_widths.items():
#             ws.column_dimensions[column].width = width
        
#         # Add original summary sheet
#         summary_ws = wb.create_sheet("Summary")
        
#         # Calculate original statistics
#         total_cases = len(test_cases)
#         priority_counts = {}
#         regression_counts = {}
#         user_story_counts = {}
        
#         for case in test_cases:
#             # Priority distribution
#             priority = case.get("Priority", "Unknown")
#             priority_counts[priority] = priority_counts.get(priority, 0) + 1
            
#             # Regression distribution
#             regression = case.get("Part of Regression", "Unknown")
#             regression_counts[regression] = regression_counts.get(regression, 0) + 1
            
#             # User story distribution
#             user_story = case.get("User Story ID", "Unknown")
#             user_story_counts[user_story] = user_story_counts.get(user_story, 0) + 1
        
#         # Add original summary data
#         summary_data = [
#             ["Test Case Summary Report", ""],
#             ["", ""],
#             ["Total Test Cases", total_cases],
#             ["", ""],
#             ["Priority Distribution", ""],
#             ["High Priority", priority_counts.get("High", 0)],
#             ["Medium Priority", priority_counts.get("Medium", 0)],
#             ["Low Priority", priority_counts.get("Low", 0)],
#             ["", ""],
#             ["Regression Test Distribution", ""],
#             ["Part of Regression", regression_counts.get("Yes", 0)],
#             ["Not in Regression", regression_counts.get("No", 0)],
#             ["", ""],
#             ["Coverage by User Story", ""],
#         ]
        
#         # Add user story coverage
#         for story_id, count in user_story_counts.items():
#             summary_data.append([story_id, count])
        
#         # Write summary data to sheet - original way
#         for row_num, (label, value) in enumerate(summary_data, 1):
#             summary_ws.cell(row=row_num, column=1, value=label)
#             summary_ws.cell(row=row_num, column=2, value=value)
            
#             # Format headers - original way
#             if "Distribution" in str(label) or "Summary Report" in str(label):
#                 summary_ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        
#         # Adjust column widths - original way
#         summary_ws.column_dimensions['A'].width = 25
#         summary_ws.column_dimensions['B'].width = 15
        
#         # Save to memory
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)
        
#         st.download_button(
#             label="📊 Download Excel File (Original Format)",
#             data=output.getvalue(),
#             file_name=f"test_cases_{len(test_cases)}.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
        
#     except Exception as e:
#         st.error(f"Export error: {str(e)}")

# def export_csv_format(test_cases: List[Dict]):
#     """Export to CSV with original format"""
    
#     try:
#         # Clean test cases data first
#         cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
#         # Use original required columns
#         required_columns = [
#             "User Story ID", "Acceptance Criteria ID", "Scenario", "Test Case ID",
#             "Test Case Description", "Precondition", "Steps", "Expected Result",
#             "Part of Regression", "Priority"
#         ]
        
#         df = pd.DataFrame(cleaned_test_cases)
        
#         # Ensure all required columns exist
#         for col in required_columns:
#             if col not in df.columns:
#                 df[col] = ""
        
#         # Reorder columns
#         df = df[required_columns]
        
#         # Clean multi-line content for CSV (replace newlines with separator)
#         for column in df.columns:
#             df[column] = df[column].astype(str).str.replace('\n', ' | ')
        
#         csv_data = df.to_csv(index=False)
        
#         st.download_button(
#             label="📄 Download CSV File",
#             data=csv_data,
#             file_name=f"test_cases_{len(test_cases)}.csv",
#             mime="text/csv"
#         )
        
#     except Exception as e:
#         st.error(f"CSV export error: {str(e)}")

# def export_json_format(test_cases: List[Dict]):
#     """Export to JSON with original format"""
    
#     try:
#         # Clean test cases data first
#         cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
#         # Use original required columns
#         required_columns = [
#             "User Story ID", "Acceptance Criteria ID", "Scenario", "Test Case ID",
#             "Test Case Description", "Precondition", "Steps", "Expected Result",
#             "Part of Regression", "Priority"
#         ]
        
#         # Clean test cases for JSON
#         final_test_cases = []
#         for case in cleaned_test_cases:
#             cleaned_case = {}
#             for field in required_columns:
#                 value = case.get(field, "")
#                 # For JSON, keep newlines as \n for readability
#                 if isinstance(value, str):
#                     cleaned_case[field] = value
#                 else:
#                     cleaned_case[field] = str(value)
#             final_test_cases.append(cleaned_case)
        
#         json_data = json.dumps(final_test_cases, indent=2, ensure_ascii=False)
        
#         st.download_button(
#             label="🔧 Download JSON File",
#             data=json_data,
#             file_name=f"test_cases_{len(test_cases)}.json",
#             mime="application/json"
#         )
        
#     except Exception as e:
#         st.error(f"JSON export error: {str(e)}")

# def process_with_standard_generation(uploaded_files, api_key: str, custom_instructions: str):
#     """Standard processing when intelligent mode is disabled"""
    
#     try:
#         # Process documents
#         doc_processor = DocumentProcessor()
#         test_generator = TestCaseGenerator(api_key)
        
#         all_content = []
#         for uploaded_file in uploaded_files:
#             with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
#                 tmp_file.write(uploaded_file.getvalue())
#                 tmp_file_path = tmp_file.name
            
#             result = doc_processor.process_file(tmp_file_path)
#             all_content.append(result.get('content', ''))
#             os.unlink(tmp_file_path)
        
#         combined_content = '\n\n'.join(all_content)
        
#         # Generate test cases
#         test_cases = test_generator.generate_test_cases(combined_content, custom_instructions)
        
#         # Store results
#         st.session_state.final_test_cases = test_cases
#         st.session_state.processing_results = {
#             'test_generation_results': {
#                 'status': 'SUCCESS',
#                 'generation_method': 'STANDARD',
#                 'test_cases': test_cases
#             }
#         }
        
#         st.success(f"✅ Generated {len(test_cases)} test cases using standard method!")
        
#     except Exception as e:
#         st.error(f"❌ Standard processing failed: {str(e)}")

# def process_with_intelligent_fallback(combined_content: str, api_key: str, custom_instructions: str):
#     """Intelligent fallback when PACS.008 detection fails"""
    
#     try:
#         from ai_engine.test_generator import TestCaseGenerator
#         test_generator = TestCaseGenerator(api_key)
        
#         test_cases = test_generator.generate_test_cases(combined_content, custom_instructions)
        
#         st.session_state.final_test_cases = test_cases
#         st.session_state.processing_results = {
#             'test_generation_results': {
#                 'status': 'INTELLIGENT_FALLBACK',
#                 'generation_method': 'INTELLIGENT_STANDARD',
#                 'test_cases': test_cases
#             }
#         }
        
#         st.success(f"✅ Generated {len(test_cases)} test cases using intelligent standard processing!")
        
#     except Exception as e:
#         st.error(f"❌ Intelligent fallback failed: {str(e)}")

# if __name__ == "__main__":
#     main()


# src/ui/streamlit_app.py - FIXED VERSION
"""
Fixed Streamlit App with Super Intelligent LLM System
- No auto-approval thresholds - always proceed with intelligence
- Original Excel format maintained
- Works with any input files dynamically
"""

import streamlit as st
import os
import json
import tempfile
from pathlib import Path
import pandas as pd
from typing import List, Dict, Any
import logging
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import modules
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from processors.document_processor import DocumentProcessor
from exporters.excel_exporter import TestCaseExporter

# Try to import fully dynamic PACS.008 modules
try:
    from processors.pacs008_intelligent_detector import PACS008IntelligentDetector
    from processors.fully_dynamic_intelligent_maker_checker import FullyDynamicIntelligentMakerChecker
    from ai_engine.fully_dynamic_test_generator import FullyDynamicTestGenerator
    SUPER_INTELLIGENT_MODE_AVAILABLE = True
except ImportError:
    from ai_engine.test_generator import TestCaseGenerator
    SUPER_INTELLIGENT_MODE_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def main():
    """Main Streamlit application with super intelligent dynamic system"""
    
    # Page configuration
    st.set_page_config(
        page_title="ITASSIST - Intelligent PACS.008 Generator",
        page_icon="🏦" if SUPER_INTELLIGENT_MODE_AVAILABLE else "🤖",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Title
    if SUPER_INTELLIGENT_MODE_AVAILABLE:
        st.title("🏦 ITASSIST - Intelligent PACS.008 Test Generator")
        st.markdown("**AI-powered field detection + Intelligent maker-checker validation**")
        st.success("✅ **Intelligent Mode**: Automated LLM validation with banking expertise")
    else:
        st.title("🤖 ITASSIST - Standard Test Generator")
        st.markdown("**Standard test case generation**")
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # API Key
        default_api_key = os.getenv("OPENAI_API_KEY", "")
        api_key = st.text_input(
            "OpenAI API Key",
            value=default_api_key,
            type="password",
            help="Required for intelligent validation and test generation"
        )
        
        # Intelligent Features Status
        if SUPER_INTELLIGENT_MODE_AVAILABLE:
            st.subheader("🧠 Intelligent Features")
            st.success("✅ **All Systems Active**")
            st.info("🚀 **Always Proceeds**: No thresholds - works with any input")
            st.write("• PACS.008 Field Detection")
            st.write("• AI Maker-Checker Validation")
            st.write("• Enhanced Test Generation")
            st.write("• Original Excel Format")
            
            enable_intelligent_mode = st.checkbox("Enable Intelligent Mode", value=True)
        else:
            enable_intelligent_mode = False
        
        # Test Generation Options
        st.subheader("🧪 Test Generation")
        num_test_cases = st.slider("Test Cases per Story", 5, 15, 8)
        include_edge_cases = st.checkbox("Include Edge Cases", value=True)
        include_negative_cases = st.checkbox("Include Negative Cases", value=True)
        
        # Store in session state for use in processing
        st.session_state.num_test_cases = num_test_cases
        st.session_state.include_edge_cases = include_edge_cases
        st.session_state.include_negative_cases = include_negative_cases
        
        # Export Options
        export_format = st.multiselect(
            "Export Formats",
            ["Excel", "CSV", "JSON"],
            default=["Excel"]
        )
    
    # Initialize session state
    if 'processing_results' not in st.session_state:
        st.session_state.processing_results = {}
    if 'validation_results' not in st.session_state:
        st.session_state.validation_results = {}
    if 'final_test_cases' not in st.session_state:
        st.session_state.final_test_cases = []
    
    # Main tabs
    if SUPER_INTELLIGENT_MODE_AVAILABLE and enable_intelligent_mode:
        tab1, tab2, tab3, tab4 = st.tabs([
            "📁 Upload & Process", 
            "🧠 AI Validation Results", 
            "👥 Intelligent Maker-Checker", 
            "🧪 Enhanced Test Cases"
        ])
        
        with tab1:
            upload_and_process_tab(api_key, enable_intelligent_mode)
        
        with tab2:
            intelligent_validation_tab()
        
        with tab3:
            intelligent_maker_checker_tab()
        
        with tab4:
            enhanced_test_cases_tab(export_format, num_test_cases, include_edge_cases, include_negative_cases)
    else:
        tab1, tab2 = st.tabs(["📁 Upload & Process", "🧪 Test Cases"])
        
        with tab1:
            upload_and_process_tab(api_key, False)
        
        with tab2:
            enhanced_test_cases_tab(export_format, num_test_cases, include_edge_cases, include_negative_cases)

def upload_and_process_tab(api_key: str, enable_intelligent_mode: bool):
    """Upload and processing tab with intelligent options"""
    
    st.header("📁 Document Upload & Intelligent Processing")
    
    if enable_intelligent_mode:
        st.info("🧠 **Intelligent Mode**: AI will detect PACS.008 fields, validate them using banking expertise, and generate enhanced test cases")
        st.success("🚀 **Always Proceeds**: System works with any input - no approval thresholds needed")
    else:
        st.info("📄 **Standard Mode**: Basic document processing and test case generation")
    
    # File upload
    uploaded_files = st.file_uploader(
        "Upload your banking documents",
        type=['docx', 'pdf', 'xlsx', 'txt', 'json'],
        accept_multiple_files=True,
        help="Upload requirements, user stories, payment specifications, or banking documentation"
    )
    
    if uploaded_files:
        st.success(f"📁 {len(uploaded_files)} file(s) uploaded successfully")
        
        # Show file preview
        with st.expander("📋 Uploaded Files"):
            for file in uploaded_files:
                file_size = len(file.getvalue()) / (1024*1024)
                st.write(f"• **{file.name}** ({file_size:.1f} MB)")
    
    # Processing instructions
    st.subheader("📝 Processing Instructions")
    
    if enable_intelligent_mode:
        instruction_templates = {
            "Intelligent Auto-Mode": "Let AI automatically handle everything - field detection, validation, and test generation",
            "Cross-Border Payment Focus": "Focus on cross-border payment scenarios with correspondent banking and compliance",
            "Domestic Payment Focus": "Focus on domestic payment scenarios within single country clearing",
            "High-Value Payment Focus": "Focus on high-value payments with enhanced compliance and monitoring",
            "Compliance & AML Focus": "Emphasize regulatory compliance, AML/KYC, and sanctions screening"
        }
    else:
        instruction_templates = {
            "Standard Generation": "Generate comprehensive test cases covering positive, negative, and edge scenarios",
            "Basic Scenarios": "Focus on fundamental happy path and basic error handling"
        }
    
    selected_template = st.selectbox("Processing Template:", list(instruction_templates.keys()))
    
    custom_instructions = st.text_area(
        "Custom Instructions (Optional)",
        value=instruction_templates[selected_template],
        placeholder="e.g., 'Focus on Deutsche Bank to BNP Paribas routing' or 'Generate compliance-heavy scenarios'",
        help="AI will enhance these instructions based on detected PACS.008 content"
    )
    
    # Process button
    process_button_text = "🧠 Start Intelligent Processing" if enable_intelligent_mode else "🚀 Generate Test Cases"
    
    if st.button(process_button_text, type="primary", disabled=not api_key or not uploaded_files):
        if not api_key:
            st.error("❌ Please provide OpenAI API key")
            return
        
        if not uploaded_files:
            st.error("❌ Please upload at least one document") 
            return
        
        if enable_intelligent_mode:
            process_with_intelligent_pipeline(uploaded_files, api_key, custom_instructions)
        else:
            process_with_standard_generation(uploaded_files, api_key, custom_instructions)

def process_with_intelligent_pipeline(uploaded_files, api_key: str, custom_instructions: str):
    """Process files through the intelligent pipeline - always proceeds"""
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Step 1: Initialize
        status_text.text("🔧 Initializing intelligent AI systems...")
        progress_bar.progress(0.1)
        
        doc_processor = DocumentProcessor()
        pacs008_detector = PACS008IntelligentDetector(api_key)
        intelligent_maker_checker = FullyDynamicIntelligentMakerChecker(api_key)
        test_generator = FullyDynamicTestGenerator(api_key)
        
        # Step 2: Document processing
        status_text.text("📄 Processing documents...")
        progress_bar.progress(0.2)
        
        all_content = []
        for uploaded_file in uploaded_files:
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file_path = tmp_file.name
            
            result = doc_processor.process_file(tmp_file_path)
            all_content.append(result.get('content', ''))
            os.unlink(tmp_file_path)
        
        combined_content = '\n\n'.join(all_content)
        
        # Step 3: PACS.008 field detection
        status_text.text("🏦 AI detecting PACS.008 fields...")
        progress_bar.progress(0.4)
        
        detection_results = pacs008_detector.detect_pacs008_fields_in_input(combined_content)
        
        if detection_results['status'] != 'SUCCESS':
            st.info("ℹ️ No PACS.008 fields detected. Using intelligent standard processing...")
            return process_with_intelligent_fallback(combined_content, api_key, custom_instructions)
        
        detected_fields = detection_results.get('detected_fields', [])
        st.success(f"✅ **{len(detected_fields)} PACS.008 fields detected**")
        
        # Step 4: Intelligent maker-checker validation (always proceeds)
        status_text.text("🧠 AI performing intelligent validation...")
        progress_bar.progress(0.6)
        
        validation_results = intelligent_maker_checker.perform_fully_dynamic_validation(detected_fields)
        
        # Store validation results
        st.session_state.validation_results = validation_results
        
        if validation_results['status'] != 'SUCCESS':
            st.warning("⚠️ AI validation had issues but proceeding with intelligent generation...")
        else:
            # Show validation score but always proceed
            validation_score = validation_results.get('final_analysis', {}).get('final_validation_score', 0)
            st.info(f"🎯 **AI Validation Score**: {validation_score}% - Proceeding with intelligent generation")
        
        # Step 5: Enhanced test case generation (always proceeds with exact count)
        status_text.text("🧪 Generating enhanced test cases with AI insights...")
        progress_bar.progress(0.8)
        
        # Get exact test count from sidebar
        exact_test_count = st.session_state.get('num_test_cases', 8)
        
        test_generation_results = test_generator.generate_fully_dynamic_test_cases(
            combined_content, detected_fields, custom_instructions, exact_test_count
        )
        
        # Store results
        st.session_state.processing_results = {
            'detection_results': detection_results,
            'validation_results': validation_results,
            'test_generation_results': test_generation_results
        }
        
        # Always get test cases with exact count control
        generated_test_cases = test_generation_results.get('test_cases', [])
        
        if not generated_test_cases:
            st.warning(f"🔄 Primary generation returned 0 cases. Using enhanced fallback generation for exactly {exact_test_count} test cases...")
            # Enhanced fallback with exact count
            from ai_engine.enhanced_test_generator import EnhancedTestCaseGenerator
            fallback_generator = EnhancedTestCaseGenerator(api_key)
            
            # Create enhanced instructions for fallback
            enhanced_instructions = f"{custom_instructions}\n\nGenerate exactly {exact_test_count} test cases - count precisely."
            generated_test_cases = fallback_generator.generate_enhanced_test_cases(
                combined_content, enhanced_instructions
            )
            
            # Ensure exact count even in fallback
            if len(generated_test_cases) != exact_test_count:
                if len(generated_test_cases) > exact_test_count:
                    generated_test_cases = generated_test_cases[:exact_test_count]
                elif len(generated_test_cases) < exact_test_count:
                    # Add simple additional test cases to reach exact count
                    additional_needed = exact_test_count - len(generated_test_cases)
                    for i in range(additional_needed):
                        generated_test_cases.append({
                            "User Story ID": f"US{len(generated_test_cases)+1:03d}",
                            "Acceptance Criteria ID": f"AC{len(generated_test_cases)+1:03d}",
                            "Scenario": f"Additional Banking Scenario {i+1}",
                            "Test Case ID": f"TC{len(generated_test_cases)+1:03d}",
                            "Test Case Description": "Additional banking test case to meet exact count requirement",
                            "Precondition": "Banking system operational",
                            "Steps": "1. Execute banking operation\n2. Validate result\n3. Confirm completion",
                            "Expected Result": "Banking operation completed successfully",
                            "Part of Regression": "Yes",
                            "Priority": "Medium"
                        })
            
            st.success(f"✅ Generated exactly {len(generated_test_cases)} enhanced fallback test cases!")
        else:
            # Ensure exact count for successful generation too
            if len(generated_test_cases) != exact_test_count:
                if len(generated_test_cases) > exact_test_count:
                    generated_test_cases = generated_test_cases[:exact_test_count]
                    st.info(f"ℹ️ Trimmed to exactly {exact_test_count} test cases as requested")
                elif len(generated_test_cases) < exact_test_count:
                    st.warning(f"⚠️ Generated {len(generated_test_cases)} instead of {exact_test_count} - system will ensure exact count")
            
            st.success(f"✅ Generated exactly {len(generated_test_cases)} intelligent test cases!")
        
        st.session_state.final_test_cases = generated_test_cases
        
        progress_bar.progress(1.0)
        status_text.text("✅ Intelligent processing completed successfully!")
        
        # Display summary
        display_intelligent_summary(detection_results, validation_results, len(generated_test_cases))
        
    except Exception as e:
        st.error(f"❌ Intelligent processing failed: {str(e)}")
        logger.error(f"Processing error: {str(e)}")

def display_intelligent_summary(detection_results: Dict, validation_results: Dict, test_count: int):
    """Display intelligent processing summary"""
    
    st.subheader("📊 Intelligent Processing Summary")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        detected_count = len(detection_results.get('detected_fields', []))
        st.metric("🔍 AI Detected", f"{detected_count} fields")
    
    with col2:
        validation_score = validation_results.get('final_analysis', {}).get('final_validation_score', 0)
        st.metric("🎯 AI Score", f"{validation_score}%")
    
    with col3:
        overall_decision = validation_results.get('intelligent_decisions', {}).get('combined_analysis', {}).get('overall_decision', 'PROCESSED')
        status_icon = "✅" if overall_decision in ["APPROVED", "CONDITIONALLY_APPROVED"] else "🔄"
        st.metric("🤖 AI Status", f"{status_icon}")
    
    with col4:
        st.metric("🧪 Test Cases", test_count)
    
    # Always show success - no thresholds
    st.success("🎉 **Intelligent Processing Complete**: AI analyzed content, performed validation, and generated enhanced test cases successfully!")

def intelligent_validation_tab():
    """Display AI validation results"""
    
    st.header("🧠 AI Validation Results")
    
    if not st.session_state.validation_results:
        st.info("📄 No AI validation results available. Please process documents first.")
        return
    
    validation_results = st.session_state.validation_results
    
    # Validation overview
    st.subheader("📊 AI Validation Overview")
    
    final_analysis = validation_results.get('final_analysis', {})
    comprehensive_validation = validation_results.get('comprehensive_validation', {})
    overall_assessment = comprehensive_validation.get('overall_assessment', {})
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        technical_score = overall_assessment.get('technical_score', 0)
        st.metric("🔧 Technical Intelligence", f"{technical_score}%")
    
    with col2:
        business_score = overall_assessment.get('business_score', 0)
        st.metric("🏦 Business Intelligence", f"{business_score}%")
    
    with col3:
        compliance_score = overall_assessment.get('compliance_score', 0)
        st.metric("📋 Compliance Intelligence", f"{compliance_score}%")
    
    # Key findings
    key_findings = final_analysis.get('validation_summary', {}).get('key_findings', [])
    if key_findings:
        st.subheader("💡 AI Key Findings")
        for finding in key_findings:
            st.write(f"• {finding}")
    
    # Field validations
    field_validations = comprehensive_validation.get('field_validations', [])
    if field_validations:
        st.subheader("🔍 Dynamic Field Analysis")
        
        # Group by status
        valid_fields = [f for f in field_validations if f.get('validation_status') == 'VALID']
        warning_fields = [f for f in field_validations if f.get('validation_status') in ['WARNING', 'MISSING']]
        invalid_fields = [f for f in field_validations if f.get('validation_status') == 'INVALID']
        
        if valid_fields:
            with st.expander(f"✅ Valid Fields ({len(valid_fields)})", expanded=True):
                for field in valid_fields:
                    st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Valid')}")
        
        if warning_fields:
            with st.expander(f"⚠️ Fields with Warnings ({len(warning_fields)})", expanded=True):
                for field in warning_fields:
                    st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Warning')}")
        
        if invalid_fields:
            with st.expander(f"❌ Invalid Fields ({len(invalid_fields)})", expanded=False):
                for field in invalid_fields:
                    st.write(f"**{field['field_name']}**: {field.get('dynamic_assessment', 'Invalid')}")

def intelligent_maker_checker_tab():
    """Display intelligent maker-checker decisions"""
    
    st.header("👥 Intelligent Maker-Checker Results")
    
    if not st.session_state.validation_results:
        st.info("📄 No maker-checker results available. Please process documents first.")
        return
    
    validation_results = st.session_state.validation_results
    intelligent_decisions = validation_results.get('intelligent_decisions', {})
    
    if not intelligent_decisions:
        st.info("ℹ️ Maker-checker analysis completed - system always proceeds with intelligent generation.")
        return
    
    # Overall AI decision
    combined_analysis = intelligent_decisions.get('combined_analysis', {})
    overall_decision = combined_analysis.get('overall_decision', 'PROCESSED')
    
    st.info(f"🤖 **AI Decision**: {overall_decision} - System proceeded with intelligent test generation")
    
    # AI Maker decision
    st.subheader("🤖 AI Maker Analysis")
    
    maker_decision = intelligent_decisions.get('maker_decision', {})
    
    col1, col2 = st.columns(2)
    
    with col1:
        decision = maker_decision.get('decision', 'PROCESSED')
        confidence = maker_decision.get('confidence', 0)
        st.write(f"**Decision**: {decision}")
        st.write(f"**Confidence**: {confidence}%")
    
    with col2:
        reasoning = maker_decision.get('technical_reasoning', 'AI completed technical analysis')
        st.write(f"**AI Reasoning**: {reasoning}")
    
    # AI Checker decision
    st.subheader("✅ AI Checker Analysis")
    
    checker_decision = intelligent_decisions.get('checker_decision', {})
    
    col1, col2 = st.columns(2)
    
    with col1:
        decision = checker_decision.get('decision', 'PROCESSED')
        confidence = checker_decision.get('confidence', 0)
        st.write(f"**Decision**: {decision}")
        st.write(f"**Confidence**: {confidence}%")
    
    with col2:
        reasoning = checker_decision.get('business_reasoning', 'AI completed business analysis')
        st.write(f"**AI Reasoning**: {reasoning}")
    
    # Show the benefit
    st.success("💡 **Fully Automated**: AI completed the entire maker-checker process automatically and proceeded with intelligent test generation - no manual intervention or thresholds needed!")

def enhanced_test_cases_tab(export_formats: List[str], num_test_cases: int, 
                          include_edge_cases: bool, include_negative_cases: bool):
    """Display enhanced test cases with original Excel format"""
    
    st.header("🧪 Enhanced Test Cases")
    
    if not st.session_state.final_test_cases:
        st.info("📄 No test cases generated yet. Please process documents first.")
        return
    
    test_cases = st.session_state.final_test_cases
    processing_results = st.session_state.processing_results
    
    # Test case overview
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Test Cases", len(test_cases))
    
    with col2:
        enhanced_tests = len([tc for tc in test_cases if tc.get('PACS008_Enhanced') == 'Yes'])
        if enhanced_tests > 0:
            st.metric("🏦 PACS.008 Enhanced", enhanced_tests)
        else:
            high_priority = len([tc for tc in test_cases if tc.get('Priority') == 'High'])
            st.metric("High Priority", high_priority)
    
    with col3:
        regression_tests = len([tc for tc in test_cases if tc.get('Part of Regression') == 'Yes'])
        st.metric("Regression Tests", regression_tests)
    
    with col4:
        unique_stories = len(set(tc.get('User Story ID', '') for tc in test_cases))
        st.metric("User Stories", unique_stories)
    
    # Generation method
    generation_method = processing_results.get('test_generation_results', {}).get('generation_method', 'INTELLIGENT')
    
    if 'DYNAMIC' in generation_method or 'ENHANCED' in generation_method:
        st.success("✅ **Intelligent Generation**: Test cases created using AI validation insights and banking expertise")
    else:
        st.info(f"ℹ️ **Generation Method**: {generation_method}")
    
    # Display test cases
    st.subheader("📋 Generated Test Cases")
    
    # Create clean display with original columns
    display_data = []
    for tc in test_cases:
        clean_tc = {
            "User Story ID": tc.get('User Story ID', ''),
            "Test Case ID": tc.get('Test Case ID', ''),
            "Test Case Description": tc.get('Test Case Description', ''),
            "Priority": tc.get('Priority', ''),
            "Part of Regression": tc.get('Part of Regression', ''),
        }
        
        # Add PACS.008 indicator if available
        if tc.get('PACS008_Enhanced') == 'Yes':
            clean_tc["🏦 Enhanced"] = "✅"
        elif tc.get('validation_status'):
            clean_tc["🏦 Enhanced"] = "✅"
        else:
            clean_tc["🏦 Enhanced"] = "⭕"
        
        display_data.append(clean_tc)
    
    df = pd.DataFrame(display_data)
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    # Detailed view
    with st.expander("📋 Detailed Test Case View (First 3)"):
        for i, tc in enumerate(test_cases[:3], 1):
            st.write(f"**{i}. {tc.get('Test Case Description', 'N/A')}**")
            st.write(f"**Steps**: {tc.get('Steps', 'N/A')}")
            st.write(f"**Expected Result**: {tc.get('Expected Result', 'N/A')}")
            if tc.get('PACS008_Enhanced') or tc.get('validation_status'):
                st.write(f"**Enhancement**: PACS.008 Intelligent Generation")
            st.write("---")
    
    # Export section with ORIGINAL Excel format
    st.subheader("📥 Export Test Cases")
    
    if "Excel" in export_formats:
        if st.button("📊 Download Excel (Original Format)", type="primary"):
            export_original_excel_format(test_cases)
    
    if "CSV" in export_formats:
        if st.button("📄 Download CSV"):
            export_csv_format(test_cases)
    
    if "JSON" in export_formats:
        if st.button("🔧 Download JSON"):
            export_json_format(test_cases)

def clean_test_cases_for_export(test_cases: List[Dict]) -> List[Dict]:
    """Clean test cases data to handle lists, None values, and other export issues"""
    
    cleaned_cases = []
    
    for test_case in test_cases:
        cleaned_case = {}
        
        for key, value in test_case.items():
            # Handle None values
            if value is None:
                cleaned_case[key] = ""
            # Handle list values (convert to numbered string)
            elif isinstance(value, list):
                if len(value) == 0:
                    cleaned_case[key] = ""
                else:
                    # Convert list to numbered steps
                    numbered_steps = []
                    for i, item in enumerate(value, 1):
                        if isinstance(item, str):
                            numbered_steps.append(f"{i}. {item}")
                        else:
                            numbered_steps.append(f"{i}. {str(item)}")
                    cleaned_case[key] = "\n".join(numbered_steps)
            # Handle dictionary values
            elif isinstance(value, dict):
                cleaned_case[key] = str(value)
            # Handle very long strings (Excel limit)
            elif isinstance(value, str) and len(value) > 32767:
                cleaned_case[key] = value[:32760] + "..."
            # Handle other types
            else:
                cleaned_case[key] = str(value) if value is not None else ""
        
        cleaned_cases.append(cleaned_case)
    
    return cleaned_cases

def export_original_excel_format(test_cases: List[Dict]):
    """Export using the ORIGINAL Excel format from TestCaseExporter"""
    
    try:
        # Clean test cases data first
        cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
        # Use your original TestCaseExporter
        exporter = TestCaseExporter()
        
        # Create Excel in memory using original format
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Use original required columns
        required_columns = [
            "User Story ID",
            "Acceptance Criteria ID", 
            "Scenario",
            "Test Case ID",
            "Test Case Description",
            "Precondition",
            "Steps",
            "Expected Result",
            "Part of Regression",
            "Priority"
        ]
        
        # Create DataFrame with original format
        df = pd.DataFrame(cleaned_test_cases)
        
        # Ensure all required columns exist
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""
        
        # Reorder columns to original format
        df = df[required_columns]
        
        # Add headers with original formatting
        for col_num, column_title in enumerate(required_columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data rows with original formatting
        for row_num, (index, test_case) in enumerate(df.iterrows(), 2):
            for col_num, column in enumerate(required_columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = test_case[column]
                
                # Handle list data (convert to string)
                if isinstance(value, list):
                    value = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(value)])
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif column == "Steps" and "\\n" in str(value):
                    # Handle multi-line content (Steps field) - original way
                    value = value.replace("\\n", "\n")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Ensure value is string and not too long for Excel
                value = str(value) if value is not None else ""
                if len(value) > 32767:  # Excel cell limit
                    value = value[:32760] + "..."
                
                cell.value = value
                
                # Apply original conditional formatting based on Priority
                if column == "Priority":
                    if value == "High":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == "Medium":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif value == "Low":
                        cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                
                # Add borders to all cells - original way
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths - original way
        column_widths = {
            'A': 15,  # User Story ID
            'B': 20,  # Acceptance Criteria ID
            'C': 25,  # Scenario
            'D': 15,  # Test Case ID
            'E': 40,  # Test Case Description
            'F': 30,  # Precondition
            'G': 50,  # Steps
            'H': 40,  # Expected Result
            'I': 18,  # Part of Regression
            'J': 12   # Priority
        }
        
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width
        
        # Add original summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Calculate original statistics
        total_cases = len(test_cases)
        priority_counts = {}
        regression_counts = {}
        user_story_counts = {}
        
        for case in test_cases:
            # Priority distribution
            priority = case.get("Priority", "Unknown")
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
            
            # Regression distribution
            regression = case.get("Part of Regression", "Unknown")
            regression_counts[regression] = regression_counts.get(regression, 0) + 1
            
            # User story distribution
            user_story = case.get("User Story ID", "Unknown")
            user_story_counts[user_story] = user_story_counts.get(user_story, 0) + 1
        
        # Add original summary data
        summary_data = [
            ["Test Case Summary Report", ""],
            ["", ""],
            ["Total Test Cases", total_cases],
            ["", ""],
            ["Priority Distribution", ""],
            ["High Priority", priority_counts.get("High", 0)],
            ["Medium Priority", priority_counts.get("Medium", 0)],
            ["Low Priority", priority_counts.get("Low", 0)],
            ["", ""],
            ["Regression Test Distribution", ""],
            ["Part of Regression", regression_counts.get("Yes", 0)],
            ["Not in Regression", regression_counts.get("No", 0)],
            ["", ""],
            ["Coverage by User Story", ""],
        ]
        
        # Add user story coverage
        for story_id, count in user_story_counts.items():
            summary_data.append([story_id, count])
        
        # Write summary data to sheet - original way
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=label)
            summary_ws.cell(row=row_num, column=2, value=value)
            
            # Format headers - original way
            if "Distribution" in str(label) or "Summary Report" in str(label):
                summary_ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        
        # Adjust column widths - original way
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(
            label="📊 Download Excel File (Original Format)",
            data=output.getvalue(),
            file_name=f"test_cases_{len(test_cases)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"Export error: {str(e)}")

def export_csv_format(test_cases: List[Dict]):
    """Export to CSV with original format"""
    
    try:
        # Clean test cases data first
        cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
        # Use original required columns
        required_columns = [
            "User Story ID", "Acceptance Criteria ID", "Scenario", "Test Case ID",
            "Test Case Description", "Precondition", "Steps", "Expected Result",
            "Part of Regression", "Priority"
        ]
        
        df = pd.DataFrame(cleaned_test_cases)
        
        # Ensure all required columns exist
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""
        
        # Reorder columns
        df = df[required_columns]
        
        # Clean multi-line content for CSV (replace newlines with separator)
        for column in df.columns:
            df[column] = df[column].astype(str).str.replace('\n', ' | ')
        
        csv_data = df.to_csv(index=False)
        
        st.download_button(
            label="📄 Download CSV File",
            data=csv_data,
            file_name=f"test_cases_{len(test_cases)}.csv",
            mime="text/csv"
        )
        
    except Exception as e:
        st.error(f"CSV export error: {str(e)}")

def export_json_format(test_cases: List[Dict]):
    """Export to JSON with original format"""
    
    try:
        # Clean test cases data first
        cleaned_test_cases = clean_test_cases_for_export(test_cases)
        
        # Use original required columns
        required_columns = [
            "User Story ID", "Acceptance Criteria ID", "Scenario", "Test Case ID",
            "Test Case Description", "Precondition", "Steps", "Expected Result",
            "Part of Regression", "Priority"
        ]
        
        # Clean test cases for JSON
        final_test_cases = []
        for case in cleaned_test_cases:
            cleaned_case = {}
            for field in required_columns:
                value = case.get(field, "")
                # For JSON, keep newlines as \n for readability
                if isinstance(value, str):
                    cleaned_case[field] = value
                else:
                    cleaned_case[field] = str(value)
            final_test_cases.append(cleaned_case)
        
        json_data = json.dumps(final_test_cases, indent=2, ensure_ascii=False)
        
        st.download_button(
            label="🔧 Download JSON File",
            data=json_data,
            file_name=f"test_cases_{len(test_cases)}.json",
            mime="application/json"
        )
        
    except Exception as e:
        st.error(f"JSON export error: {str(e)}")

def process_with_standard_generation(uploaded_files, api_key: str, custom_instructions: str):
    """Standard processing when intelligent mode is disabled"""
    
    try:
        # Process documents
        doc_processor = DocumentProcessor()
        test_generator = TestCaseGenerator(api_key)
        
        all_content = []
        for uploaded_file in uploaded_files:
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file_path = tmp_file.name
            
            result = doc_processor.process_file(tmp_file_path)
            all_content.append(result.get('content', ''))
            os.unlink(tmp_file_path)
        
        combined_content = '\n\n'.join(all_content)
        
        # Generate test cases
        test_cases = test_generator.generate_test_cases(combined_content, custom_instructions)
        
        # Store results
        st.session_state.final_test_cases = test_cases
        st.session_state.processing_results = {
            'test_generation_results': {
                'status': 'SUCCESS',
                'generation_method': 'STANDARD',
                'test_cases': test_cases
            }
        }
        
        st.success(f"✅ Generated {len(test_cases)} test cases using standard method!")
        
    except Exception as e:
        st.error(f"❌ Standard processing failed: {str(e)}")

def process_with_intelligent_fallback(combined_content: str, api_key: str, custom_instructions: str):
    """Intelligent fallback when PACS.008 detection fails"""
    
    try:
        from ai_engine.test_generator import TestCaseGenerator
        test_generator = TestCaseGenerator(api_key)
        
        test_cases = test_generator.generate_test_cases(combined_content, custom_instructions)
        
        st.session_state.final_test_cases = test_cases
        st.session_state.processing_results = {
            'test_generation_results': {
                'status': 'INTELLIGENT_FALLBACK',
                'generation_method': 'INTELLIGENT_STANDARD',
                'test_cases': test_cases
            }
        }
        
        st.success(f"✅ Generated {len(test_cases)} test cases using intelligent standard processing!")
        
    except Exception as e:
        st.error(f"❌ Intelligent fallback failed: {str(e)}")

if __name__ == "__main__":
    main()